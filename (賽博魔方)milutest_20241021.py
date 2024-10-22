# 套件匯入
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pyautogui as pag
from selenium.webdriver.chrome.service import Service
import requests
from http import HTTPStatus
from selenium.webdriver.chrome.options import Options
from lxml import html
import csv
import numpy as np
import pandas as pd
import cv2
import pybi as pbi
import os
import sys
import aspose
import xlrd
import xlwt
import glob
import cx_Freeze
from cx_Freeze import setup
import setuptools
import jpype
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType
import difflib
import openpyxl
from openpyxl import Workbook
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl.styles import Font  # 導入字體模組
from openpyxl.styles import PatternFill  # 導入填充模組
from spire.xls import *
from collections import deque
from pandas.core.frame import DataFrame
# 导入openpyxl模块并将其重命名为pxl
import openpyxl as pxl
# 从openpyxl导入PatternFill类
from openpyxl.styles import PatternFill
# 从openpyxl导入colors类
from openpyxl.styles import colors
# 从openpyxl导入Font类
from openpyxl.styles import Font
import datetime
import pytesseract
from PIL import Image
import ddddocr


current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
print('測試起始時間: ', current_time, '\n')


print('串接(賽博魔方)前台API擷取遊戲紀錄測試中...', '\n')
time.sleep(1)

ch_options = Options()
ch_options.add_argument("--headless")  # 無WEB UI顯示
driver = webdriver.Chrome(ch_options)
   
# driver = webdriver.Chrome() 
driver.get('https://dev-api.iplaystar.net/gamehistory/?host_id=dd12f3635bc9599ca58012601d07197c&uid=ufd37xsOxWxKzc%2Bs6aWgwtLKRv%2BSBW9SwmGkHwEGpfA%3D&page=1&count=20&type=3&category=SLOT&lang=sch&game_id=PSS-ON-00158&sel_date=2024-10-21')
http_status = requests.get('https://dev-api.iplaystar.net/gamehistory/?host_id=dd12f3635bc9599ca58012601d07197c&uid=ufd37xsOxWxKzc%2Bs6aWgwtLKRv%2BSBW9SwmGkHwEGpfA%3D&page=1&count=20&type=3&category=SLOT&lang=sch&game_id=PSS-ON-00158&sel_date=2024-10-21')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')
time.sleep(2)

# # =================================選取記錄日期 =================================

# dateList = WebDriverWait(driver, 5).until(
#     EC.element_to_be_clickable((By.XPATH, '//*[@id="serdate"]')))  # 點選日期選單
# actions = ActionChains(driver)
# actions.move_to_element(dateList)    
# actions.perform()
# dateList.click()
# time.sleep(2)

# # recordDate = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/thead/tr[2]/th[1]')  # 選取 "<"鍵跳至前月日期選單
# # actions.move_to_element(recordDate)    
# # actions.perform()
# # recordDate.click()

# date_select = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/tbody/tr[3]/td[5]')  # 遊戲紀錄日期(2024.10.21)
# date_select.click()
# time.sleep(1)


# =======================================================寫入第一筆測試紀錄(一般遊戲)=======================================================

xpath1 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]')
textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newDataList1 = DataFrame(newList1)
newList1T = newDataList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021).csv', index=0, encoding='utf-8-sig')
readCSV = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021).csv')

newData = readCSV.rename(columns = {"0" : '代理商D', "1" : '遊戲類型D', "2" : '紀錄流水號', "3" : '代理伺服器', "4" : '玩家名稱', 
                                     "5" :'遊戲名稱', "6" :'遊戲時間D', "7" : '場景D', "8" : '遊戲日期D', "9" : '遊戲時間', 
                                     "10" : '場景', "11" : '序號D', "12" : '押注D', "13" : '序號', "14" : '押注', 
                                     "15" : '彩金D', "16" : '贏分D', "17" : '彩金', "18" : '贏分'})

csvFile1 = newData.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v1.csv', index=None, encoding='utf-8-sig')
print("第", 1, "筆資料寫入完成。", '\n')
time.sleep(1)

for i in range(2, 6):
    xpath2 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath2)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')


# =========================================(免遊)遊戲紀錄(無特色)=========================================      

xpath1F = driver.find_element(By.XPATH, '//*[@id="content"]/div[6]')
textSplit1F = xpath1F.text.split()
newList1F = list(textSplit1F)
newDataList1F = DataFrame(newList1F)
newList1FT = newDataList1F.T
csvFile = newList1FT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021).csv', index=0, encoding='utf-8-sig')
readCSV = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021).csv')

newData = readCSV.rename(columns = {"0" : '代理商D', "1" : '遊戲類型D', "2" : '紀錄流水號', "3" :'代理伺服器', "4" : '玩家名稱',
                                    "5" : '遊戲名稱', "6" : '遊戲時間D', "7" : '場景D', "8" : '遊戲日期D', "9" : '遊戲時間',
                                    "10" : '場景', "11" : '序號D', "12" : '押注D', "13" : '序號', "14" : '押注',
                                    "15" : '贏分D', "16" : "贏分"})

csvFile1 = newData.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v1.csv', index=None, encoding='utf-8-sig')
print("第", 6, "筆資料寫入完成。", '\n')
time.sleep(1)

for i in range(7, 19):
    xpath2F = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath2F)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')


# =========================================免遊(前)一筆遊戲紀錄(無特色)=========================================

xpath2 = driver.find_element(By.XPATH, '//*[@id="content"]/div[19]')
textSplit2 = xpath2.text.split()
newList2 = list(textSplit2)
newDataList2 = DataFrame(newList2)
newList2 = newDataList2.T
csvFile = newList2.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021).csv', index=0, encoding='utf-8-sig')
readCSV = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021).csv')

newData = readCSV.rename(columns = {"0" : '代理商D', "1" : '遊戲類型D', "2" : '紀錄流水號', "3" :'代理伺服器', "4" : '玩家名稱',
                                    "5" : '遊戲名稱', "6" : '遊戲時間D', "7" : '場景D', "8" : '遊戲日期D', "9" : '遊戲時間', "10" : '場景',
                                    "11" : '序號D', "12" : '押注D', "13" : '序號', "14" : '押注', "15" : '彩金D',
                                    "16" : '贏分D', "17" : "免遊贏分D", "18" : "彩金", "19" : "贏分", "20" : "免遊贏分"})

csvFile1 = newData.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v1.csv', index=None, encoding='utf-8-sig')
print("第", 19, "筆資料寫入完成。", '\n')
time.sleep(1)


# ========================================(一般)遊戲紀錄=========================================  

for i in range(20, 34):
    xpath3 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath3)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')


# =================================選取遊戲紀錄頁(2) =================================    
# ==================================(一般)遊戲紀錄==================================  
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[2]')
page_select.click()
time.sleep(3)


for i in range(1, 21):
    xpath3 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath3)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+33, '筆資料寫入完成。', '\n')


# =================================選取遊戲紀錄頁(3) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[3]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+53, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(4) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[4]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+73, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(5) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[5]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+93, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(6) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[6]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+113, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(7) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[7]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+133, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(8) =================================    
# ==================================(免遊)遊戲紀錄==================================
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[8]')
page_select.click()
time.sleep(3)


for i in range(1, 14):
    xpath2F = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath2F)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+153, '筆資料寫入完成。', '\n')
    

# =========================================免遊(前)一筆遊戲紀錄(無特色)=========================================

xpath2 = driver.find_element(By.XPATH, '//*[@id="content"]/div[14]')
textSplit2 = xpath2.text.split()
newList2 = list(textSplit2)
newDataList2 = DataFrame(newList2)
newList2 = newDataList2.T
csvFile = newList2.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
print("第", 167, "筆資料寫入完成。", '\n')
time.sleep(1)

# ==================================(一般)遊戲紀錄================================== 

for i in range(15, 34):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+153, '筆資料寫入完成。', '\n')


# =================================選取遊戲紀錄頁(9) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[9]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+186, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(10) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[10]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+206, '筆資料寫入完成。', '\n')


# =================================選取遊戲紀錄頁(11) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[11]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+226, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(12) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[12]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+246, '筆資料寫入完成。', '\n')
    
# =================================選取遊戲紀錄頁(13) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[13]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+266, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(14) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[14]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+286, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(15) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[15]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+306, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(16) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[16]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+326, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(17) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[17]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+346, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(18) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[18]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+366, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(19) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[19]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+386, '筆資料寫入完成。', '\n')   
    

# =================================選取遊戲紀錄頁(20) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[20]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+406, '筆資料寫入完成。', '\n')   
    
    
# =================================選取遊戲紀錄頁(21) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[21]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+426, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(22) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[22]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+446, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(23) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[23]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+466, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(24) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[24]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+486, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(25) =================================    
# ==================================(一般)遊戲紀錄================================== 
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[25]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath4 = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath4)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_Record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+506, '筆資料寫入完成。', '\n')     


totalList = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v1.xlsx', index=False, header=True)

totalList = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v1.xlsx', index=False, header=True)

totalList = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v1.xlsx', index=False, header=True)

print("EXCEL檔案儲存成功!", '\n')
# time.sleep(2)   
# driver.close()
print('(賽博魔方)前台526筆遊戲紀錄獲取完成!', '\n')


# =======================================================(前台)資料處理=======================================================

print('(前台)遊戲紀錄格式化...', '\n')


excelUpdate = pd.read_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v1.xlsx').drop(
{'代理商D', '代理伺服器','遊戲類型D', '遊戲時間D', '場景D', '遊戲日期D', '序號D', "押注D", '彩金D', '贏分D'}, axis=1)  # 刪除多餘欄位
excelWrite = excelUpdate.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v2.xlsx', index=None, header=True)
csvWrite = excelUpdate.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v2.csv', header=True, index=None, encoding='utf-8-sig')

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v2.xlsx')
ws = workbook_1["Sheet1"]

ws.insert_cols(2)#插入列

workbook_1.save(r"C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v2.xlsx")


workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v2.xlsx')
source_sheet = workbook['Sheet1']

source_sheet.move_range('D1' ":" 'D499', rows = 0, cols = -2)
source_sheet.delete_cols(4)


workbook.save(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v2.xlsx')
workbook.close()


excelUpdate = pd.read_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v1.xlsx').drop(
{'代理商D', '遊戲類型D', '代理伺服器', '遊戲時間D', '場景D', '遊戲日期D', '序號D', "押注D", '贏分D'}, axis=1)  # 刪除多餘欄位
excelWrite = excelUpdate.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v2.xlsx', index=None, header=True)
csvWrite = excelUpdate.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v2.csv', header=True, index=None, encoding='utf-8-sig')

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v2.xlsx')
ws = workbook_1["Sheet1"]

ws.insert_cols(2)#插入列

workbook_1.save(r"C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v2.xlsx")

workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v2.xlsx')
source_sheet = workbook['Sheet1']

source_sheet.move_range('D1' ":" 'D27', rows = 0, cols = -2)
source_sheet.delete_cols(4)

workbook.save(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v2.xlsx')
workbook.close()


excelUpdate = pd.read_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v1.xlsx').drop(
{'代理商D', '遊戲類型D', '遊戲時間D', '場景D', '遊戲日期D', '序號D',  "彩金D", "押注D", '贏分D', '免遊贏分D'}, axis=1)  # 刪除多餘欄位
excelWrite = excelUpdate.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v2.xlsx', index=None, header=True)
csvWrite = excelUpdate.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v2.csv', header=True, index=None, encoding='utf-8-sig')

workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v2.xlsx')
source_sheet = workbook['Sheet1']

source_sheet.move_range('D1' ":" 'D3', rows = 0, cols = -2)
source_sheet.delete_cols(4)

workbook.save(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v2.xlsx')
workbook.close()


print('(前台)遊戲紀錄格式化完成!', '\n')
time.sleep(1)
driver.close()


# ============================Section.1 進入後台首頁============================

ch_options = Options()
ch_options.add_argument("--headless")  # => 無WEB UI顯示
driver = webdriver.Chrome(ch_options)

# driver = webdriver.Chrome()  # 有WEB UI顯示

print('串接DEV後台擷取遊戲紀錄測試中...', '\n')
time.sleep(1) 
driver.get("https://dev-admin-br-02.claretfox.com/")
http_status = requests.get('https://dev-admin-br-02.claretfox.com/')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print('Http Response Code:', stateCode, '\n')
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')
time.sleep(1)

back_platform = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div/div/div/form/div/div[1]")))
back_platform.click()
time.sleep(3)
print("進入DEV後台首頁!", '\n')

# -----------切換網頁顯示語系-----------
language_bar = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/form/label')))
language_bar.click()

language_ch_zh = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="lang"]/option[2]')))
language_ch_zh.click()
print("語系已切換'繁體中文'!", '\n')

# ============================Section.2 登入使用者名稱及密碼============================

login_icon = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, 'content-group')))
login_icon.click()  # 尋找登入介面元素位址
time.sleep(2)

login_acc = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="user_id"]')))
login_acc.clear()    # 預設此欄位為null, 但仍先清除帳號欄位資訊
login_pass = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]')))   
login_pass.clear()   # 預設此欄位為null, 但仍先清除密碼欄位資訊

login_verificationCode = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input')))  # 驗證碼欄位

verificationCode = driver.find_element(By.CSS_SELECTOR, '#captcha_img > img')
verificationCode.screenshot('verificationCode.png')
time.sleep(1)

actions = ActionChains(driver)
actions.move_to_element(login_acc)    
actions.perform()
login_acc.send_keys("ivan_li")   # 個人使用者帳號
time.sleep(1)
login_acc.send_keys(Keys.TAB)   # 切換至密碼輸入欄位
login_pass.send_keys("iPlaystar296")  # 個人密碼
time.sleep(1)
login_pass.send_keys(Keys.TAB)   # 切換至驗證碼輸入欄位
verifiCode = driver.find_element(By.XPATH, '//*[@id="captcha_img"]/img')  # 驗證碼圖片元素位置
data = verifiCode.screenshot_as_png
ocr = ddddocr.DdddOcr()
# with open('verificationCode.png', 'rb') as fp:
#     image = fp.read()
catch = ocr.classification(data)
login_verificationCode.send_keys(catch)
time.sleep(1)

try:
    login_button = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[6]/button')))
    login_button.click()
    time.sleep(1)
except Exception as e:
    # errorInfo = driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[7]')
    # # print('Error message = NoSuchElementException', '\n')
    # for errorCode in errorInfo:
        # print(errorCode, '\n')
    # 處理異常
    print("驗證碼輸入錯誤，請再試一次!", e)
    # print("驗證碼輸入錯誤，請再試一次!", '\n')
    driver.quit()  
else:
    print("登入成功!", '\n')

time.sleep(5)

# ============================Section.3 切換後台功能頁籤============================

player = driver.find_element(By.ID, 'Player')
player.click()
print("進入玩家功能選單!", '\n')
time.sleep(3)

driver.get('https://dev-admin-br-02.claretfox.com/Player/game_history')
print("切換遊戲紀錄子選單!", '\n')
time.sleep(2)

start_time = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div')))

startTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button')
actions.move_to_element(startTime_button)    
actions.perform()
startTime_button.click()

# startDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
# actions.move_to_element(startDate)    
# actions.perform()
# startDate.click()

startTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[1]/a')  # 開始日期選取 "2024.10.21"
startTime_check.click()
time.sleep(1)

startTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
startTime_confirm.click()
time.sleep(2)


endTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i')
actions.move_to_element(endTime_button)    
actions.perform()
endTime_button.click()

# endDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
# actions.move_to_element(endDate)    
# actions.perform()
# endDate.click()

endTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[1]/a')  # 結束日期選取 "2024.10.17"
endTime_check.click()
time.sleep(1)

endTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
endTime_confirm.click()
time.sleep(2)

showLimit = driver.find_element(By.XPATH, '//*[@id="count"]')
showLimit.click()
time.sleep(2)

showLimitSet = driver.find_element(By.XPATH, '//*[@id="count"]/option[5]')  # 設定顯示筆數 = 500筆
showLimitSet.click()
time.sleep(2)

agencyType = driver.find_element(By.XPATH, '//*[@id="agent_attr"]')
agencyType.click()
time.sleep(2)

agencyTypeSet = driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[1]')  # 設定代理商類別 = 全選
agencyTypeSet.click()
time.sleep(2)
print('代理商類別: ', agencyTypeSet.text, '\n')

agencyName = driver.find_element(By.XPATH, '//*[@id="agent"]')
agencyName.click()
time.sleep(2)

agencyNameSet = driver.find_element(By.XPATH, '//*[@id="agent"]/option[1]')  # 設定代理商名稱 = PLAYSTAR
agencyNameSet.click()
time.sleep(2)
print('代理商名稱: ', agencyNameSet.text, '\n')

gameType = driver.find_element(By.XPATH, '//*[@id="game_type"]')
gameType.click()
time.sleep(2)

gameTypeSet = driver.find_element(By.XPATH, '//*[@id="game_type"]/option[2]')  # 設定遊戲類別 = 老虎機
driver.execute_script("arguments[0].scrollIntoView()", gameTypeSet)
gameTypeSet.click()
time.sleep(2)
print('遊戲類別: ', gameTypeSet.text, '\n')

gameSelect = driver.find_element(By.XPATH, '//*[@id="game_select"]')
gameSelect.click()
time.sleep(2)

gameSelectSet = driver.find_element(By.XPATH, '//*[@id="game_select"]/option[160]')  # 遊戲選擇 = 賽博魔方
time.sleep(2)
print('遊戲名稱: ', gameSelectSet.text, '\n')

playerName = driver.find_element(By.XPATH, '//*[@id="player"]')
playerName.send_keys("milu7")
time.sleep(2)

btnSubmit = driver.find_element(By.XPATH, '//*[@id="sh_btn"]')  # 確認無誤送出
btnSubmit.click()
time.sleep(10)
print('獲取(賽博魔方)後台遊戲紀錄中...', '\n')


# =======================================================寫入第一筆測試紀錄=======================================================

xpath1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]')
textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newDataList1 = DataFrame(newList1)
newList1T = newDataList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021).csv', index=0, encoding='utf-8-sig')
readCSV = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021).csv')

newData = readCSV.rename(columns = {"0" : '紀錄流水號', "1" : '遊戲名稱', "2" : '遊戲編號', "3" : '玩家名稱D', "4" : '代理伺服器D', 
                                     "5" :'玩家名稱', "6" :'起始時間D', "7" : '結束時間D', "8" : '遊戲日期D', "9" : '起始時間D1', 
                                     "10" : '遊戲日期D1', "11" : '遊戲時間', "12" : '遊戲狀態D', "13" : '場景D', "14" : '遊戲狀態D1', 
                                     "15" : '場景', "16" : '序號D', "17" : '面額D', "18" : '帳戶D', "19" : '押注D',
                                     "20" : '序號', "21" : '面額D1', "22" : '帳戶D1', "23" : '押注', "24" : '彩金D',
                                     "25" : '贏分D', "26" : '彩金', "27" : '贏分'})

csvFile1 = newData.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.csv', index=None, encoding='utf-8-sig')
print("第", 1, "筆資料寫入完成。", '\n')
time.sleep(1)

for i in range(2, 6):
    xpath3 = f'//*[@id="history"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath3)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')


# =========================================(免遊)遊戲紀錄(無特色)=========================================      

xpath1F = driver.find_element(By.XPATH, '//*[@id="history"]/div[6]')
textSplit1F = xpath1F.text.split()
newList1F = list(textSplit1F)
newDataList1F = DataFrame(newList1F)
newList1FT = newDataList1F.T
csvFile = newList1FT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021).csv', index=0, encoding='utf-8-sig')
readCSV = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021).csv')

newData = readCSV.rename(columns = {"0" : '紀錄流水號', "1" : '遊戲名稱', "2" : '遊戲編號', "3" :'玩家D', "4" : '代理伺服器D',
                                    "5" : '玩家名稱', "6" : '遊戲時間D', "7" : '遊戲日期D', "8" : '遊戲時間', "9" : '遊戲狀態D',
                                    "10" : '場景D', "11" : '遊戲狀態D1', "12" : '場景', "13" : '序號D', "14" : '面額D',
                                    "15" : '帳戶D', "16" : "押注D", "17" : "序號", "18" : "面額", "19" : "帳戶",
                                    "20" : "押注", "21" : "贏分D", "22" : "贏分"})

csvFile1 = newData.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v1.csv', index=None, encoding='utf-8-sig')
print("第", 6, "筆資料寫入完成。", '\n')
time.sleep(1)

for i in range(7, 19):
    xpath2F = f'//*[@id="history"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath2F)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')


# =========================================免遊(前)一筆遊戲紀錄(無特色)=========================================

xpath2 = driver.find_element(By.XPATH, '//*[@id="history"]/div[19]')
textSplit2 = xpath2.text.split()
newList2 = list(textSplit2)
newDataList2 = DataFrame(newList2)
newList2 = newDataList2.T
csvFile = newList2.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021).csv', index=0, encoding='utf-8-sig')
readCSV = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021).csv')

newData = readCSV.rename(columns = {"0" : '紀錄流水號', "1" : '遊戲名稱', "2" : '遊戲編號', "3" :'玩家D', "4" : '代理伺服器D',
                                    "5" : '玩家名稱', "6" : '起始時間D', "7" : '結束時間D', "8" : '起始日期D', "9" : '起始時間D1',
                                    "10" : '結束日期D', "11" : '遊戲時間', "12" : "遊戲狀態D", "13" : "場景D", "14" : "遊戲狀態D1",
                                    "15" : "場景", "16" : "序號D", "17" : "面額D", "18" : "帳戶D", "19" : "押注D",
                                    "20" : "序號", "21" : "面額D1", "22" : "帳戶D1", "23" : "押注", "24" : "彩金D",
                                    "25" : "贏分D", "26" : "免遊贏分D", "27" : "彩金", "28" : "贏分", "29" : "免遊贏分"})

csvFile1 = newData.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v1.csv', index=None, encoding='utf-8-sig')
print("第", 19, "筆資料寫入完成。", '\n')
time.sleep(1)


# ========================================(一般)遊戲紀錄=========================================  

for i in range(20, 154):
    xpath3 = f'//*[@id="history"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath3)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')

# =========================================(免遊)遊戲紀錄(無特色)=========================================    

for i in range(154, 167):
    xpath2F = f'//*[@id="history"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath2F)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')


# =========================================免遊(前)一筆遊戲紀錄(無特色)=========================================

xpath2 = driver.find_element(By.XPATH, '//*[@id="history"]/div[167]')
textSplit2 = xpath2.text.split()
newList2 = list(textSplit2)
newDataList2 = DataFrame(newList2)
newList2 = newDataList2.T
csvFile = newList2.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
print("第", 167, "筆資料寫入完成。", '\n')
time.sleep(1)


# ========================================(一般)遊戲紀錄=========================================  

for i in range(168, 234):
    xpath3 = f'//*[@id="history"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath3)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')

for i in range(234, 235):
    xpath3 = f'//*[@id="history"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath3)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList.remove('活動')
    newDataList.remove('20.00')
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')
    
for i in range(235, 527):
    xpath3 = f'//*[@id="history"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath3)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')
    

totalList = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.xlsx', index=False, header=True)

totalList = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v1.xlsx', index=False, header=True)

totalList = pd.read_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v1.xlsx', index=False, header=True)


print("EXCEL檔案儲存成功!", '\n')
time.sleep(2)   
# driver.close()
print('(賽博魔方)後台526筆遊戲紀錄獲取完成!', '\n')


# =======================================================(後台)資料處理=======================================================

print('(後台)遊戲紀錄格式化...', '\n')


excelUpdate = pd.read_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.xlsx').drop(
{'玩家名稱D', '代理伺服器D', '遊戲編號', '起始時間D', '結束時間D', '遊戲日期D', '遊戲日期D', "起始時間D1", '遊戲日期D1',
 '遊戲狀態D', '場景D', '遊戲狀態D1', '序號D', '面額D', '帳戶D', '押注D', '面額D1', '帳戶D1','彩金D', '贏分D'}, axis=1)  # 刪除多餘欄位
excelWrite = excelUpdate.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v2.xlsx', index=None, header=True)
csvWrite = excelUpdate.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v2.csv', header=True, index=None, encoding='utf-8-sig')


excelUpdate = pd.read_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v1.xlsx').drop(
{'玩家D', '遊戲編號', '代理伺服器D', '遊戲時間D', '遊戲日期D', "遊戲狀態D", '場景D', 
 '遊戲狀態D1', '序號D', '面額D', '帳戶D', '帳戶', '押注D', '面額', '贏分D'}, axis=1)  # 刪除多餘欄位
excelWrite = excelUpdate.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v2.xlsx', index=None, header=True)
csvWrite = excelUpdate.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v2.csv', header=True, index=None, encoding='utf-8-sig')


excelUpdate = pd.read_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v1.xlsx').drop(
{'遊戲編號', '玩家D', '代理伺服器D', '起始時間D', '結束時間D', '起始日期D',
 '起始時間D1', "結束日期D", "遊戲狀態D", '場景D', '遊戲狀態D1', '序號D',
 '面額D', '帳戶D', '押注D', '面額D1', '帳戶D1', '彩金D',
 '贏分D', '免遊贏分D'}, axis=1)  # 刪除多餘欄位
excelWrite = excelUpdate.to_excel(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v2.xlsx', index=None, header=True)
csvWrite = excelUpdate.to_csv(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v2.csv', header=True, index=None, encoding='utf-8-sig')


print('(後台)遊戲紀錄格式化完成!', '\n')
time.sleep(1)


# =================================================== Step.3 前/後台資料分析 ===================================================

print('前/後台遊戲紀錄比對中...', '\n')

# 1.數據比對(一般遊戲)

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v2.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v2.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_record(milu7_20241021)_analysis.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_record(milu7_20241021)_analysis.xlsx')
workbook_2.close()
time.sleep(1)
print('前/後台遊戲紀錄比對完成!', '\n')

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v2.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v2.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_record(milu7_20241021)_analysis.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_record(milu7_20241021)_analysis.xlsx')

print('前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(General).xlsx")
print("前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = "C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(General).xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '前台原始遊戲紀錄(賽博魔方)_20241021'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '後台原始遊戲紀錄(賽博魔方)_20241021'  # 修改分頁.2工作表名稱
ws_2 = wb['Sheet1_2']
ws_2.title = '前台遊戲紀錄比對結果(賽博魔方)_20241021'  # 修改分頁.3工作表名稱
ws_3 = wb['Sheet1_3']
ws_3.title = '後台遊戲紀錄比對結果(賽博魔方)_20241021'  # 修改分頁.4工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(General).xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
sheet_1 = wb.worksheets[2]  # 分頁.3
sheet_1.sheet_properties.tabColor = 'CC6600'
sheet_2 = wb.worksheets[3]  # 分頁.3
sheet_2.sheet_properties.tabColor = 'AAFFEE'

wb.save("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(General).xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()



# 1.數據比對(免費遊戲)

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v2.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v2.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_免遊(milu7_20241021)_analysis.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_免遊(milu7_20241021)_analysis.xlsx')
workbook_2.close()
time.sleep(1)
print('前/後台遊戲紀錄比對完成!', '\n')

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v2.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v2.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_免遊(milu7_20241021)_analysis.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_免遊(milu7_20241021)_analysis.xlsx')

print('前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(免遊).xlsx")
print("前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = "C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(免遊).xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '(免遊)前台原始紀錄(賽博魔方)_20241021'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '(免遊)後台原始紀錄(賽博魔方)_20241021'  # 修改分頁.2工作表名稱
ws_2 = wb['Sheet1_2']
ws_2.title = '(免遊)前台紀錄比對結果(賽博魔方)_20241021'  # 修改分頁.3工作表名稱
ws_3 = wb['Sheet1_3']
ws_3.title = '(免遊)後台紀錄比對結果(賽博魔方)_20241021'  # 修改分頁.4工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(免遊).xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
sheet_1 = wb.worksheets[2]  # 分頁.3
sheet_1.sheet_properties.tabColor = 'CC6600'
sheet_2 = wb.worksheets[3]  # 分頁.3
sheet_2.sheet_properties.tabColor = 'AAFFEE'

wb.save("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(免遊).xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()



# 1.數據比對(免費遊戲前一筆)

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v2.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v2.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_免遊前一筆(milu7_20241021)_analysis.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_免遊前一筆(milu7_20241021)_analysis.xlsx')
workbook_2.close()
time.sleep(1)
print('前/後台遊戲紀錄比對完成!', '\n')

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v2.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v2.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_免遊前一筆(milu7_20241021)_analysis.xlsx')
files.append(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_免遊前一筆(milu7_20241021)_analysis.xlsx')

print('前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(免遊前一筆).xlsx")
print("前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = "C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(免遊前一筆).xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '(免遊前一筆)前台原始紀錄(賽博魔方)_20241021'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '(免遊前一筆)後台原始紀錄(賽博魔方)_20241021'  # 修改分頁.2工作表名稱
ws_2 = wb['Sheet1_2']
ws_2.title = '(免遊前一筆)前台紀錄比對結果(賽博魔方)_20241021'  # 修改分頁.3工作表名稱
ws_3 = wb['Sheet1_3']
ws_3.title = '(免遊前一筆)後台紀錄比對結果(賽博魔方)_20241021'  # 修改分頁.4工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(免遊前一筆).xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
sheet_1 = wb.worksheets[2]  # 分頁.3
sheet_1.sheet_properties.tabColor = 'CC6600'
sheet_2 = wb.worksheets[3]  # 分頁.3
sheet_2.sheet_properties.tabColor = 'AAFFEE'

wb.save("C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(3). Data_Merge\Data Merge_20241021(免遊前一筆).xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()

# ================================== 刪除多餘檔案 ==================================

os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021).csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v1.csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(milu7_20241021)_v1.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021).csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v1.csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊(milu7_20241021)_v1.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021).csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v1.csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_免遊前一筆(milu7_20241021)_v1.xlsx')

os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021).csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(milu7_20241021)_v1.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021).csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v1.csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊(milu7_20241021)_v1.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021).csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v1.csv')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_免遊前一筆(milu7_20241021)_v1.xlsx')

os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_record(milu7_20241021)_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_免遊(milu7_20241021)_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(1). Front-platform\front-platform_免遊前一筆(milu7_20241021)_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_record(milu7_20241021)_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_免遊(milu7_20241021)_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(4). Cyber Cube\(3). Data Analysis\(2). Back-platform\back-platform_免遊前一筆(milu7_20241021)_analysis.xlsx')


end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
print('測試結束時間: ', end_time, '\n')

time.sleep(1)
driver.close()