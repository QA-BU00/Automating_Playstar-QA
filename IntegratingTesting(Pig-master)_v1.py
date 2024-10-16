# 套件匯入
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pyautogui as pag
from selenium.webdriver.chrome.service import Service
import requests
from http import HTTPStatus
from selenium.webdriver.chrome.options import Options
from lxml import html
import csv
import numpy as np
import pandas as pd
import cv2
import pybi as pbi
import os
import sys

import aspose
import xlrd
import xlwt
import glob
from cx_Freeze import setup
import setuptools
import jpype
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType
import difflib
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font  # 導入字體模組
from openpyxl.styles import PatternFill  # 導入填充模組
from spire.xls import *
from collections import deque
from pandas.core.frame import DataFrame
# 导入openpyxl模块并将其重命名为pxl
import openpyxl as pxl
# 从openpyxl导入PatternFill类
from openpyxl.styles import PatternFill
# 从openpyxl导入colors类
from openpyxl.styles import colors
# 从openpyxl导入Font类
from openpyxl.styles import Font
import datetime
import pytesseract
from PIL import Image
import ddddocr

current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
print('測試起始時間: ', current_time, '\n')


print('串接(小豬大師)前台API擷取遊戲紀錄測試中...', '\n')
time.sleep(1)

# ch_options = Options()
# ch_options.add_argument("--headless")  # 無WEB UI顯示
# driver = webdriver.Chrome(ch_options)
   
driver = webdriver.Chrome() 
driver.get('https://dev-api.iplaystar.net/gamehistory/?host_id=aa62ffb88b300f6be6654615f17ce6fa&lang=tch&game_id=PSS-ON-00157&count=20&page=1&uid=uN22JyloR1aD2zgUIu8nL1ogWFBPqrujwHHpugHyw94%3d')
http_status = requests.get('https://dev-api.iplaystar.net/gamehistory/?host_id=aa62ffb88b300f6be6654615f17ce6fa&lang=tch&game_id=PSS-ON-00157&count=20&page=1&uid=uN22JyloR1aD2zgUIu8nL1ogWFBPqrujwHHpugHyw94%3d')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')

def errorState(errorInfo):
    errorInfo = driver.find_element(By.XPATH, '//*[@id="sf-resetcontent"]/h2/span[3]')
    if errorInfo == True:
        raise Exception("Invalid argument supplied for foreach!", errorInfo, '\n')
        # 触发异常后，后面的代码就不会再执行
time.sleep(2)                


# =================================選取記錄日期 =================================

dateList = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="serdate"]')))
actions = ActionChains(driver)
actions.move_to_element(dateList)    
actions.perform()
dateList.click()
time.sleep(2)

recordDate = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/thead/tr[2]/th[1]')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(recordDate)    
actions.perform()
recordDate.click()

date_select01 = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/thead/tr[2]/th[1]')
date_select01.click()
time.sleep(1)

date_select = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[7]/div[1]/table/tbody/tr[5]/td[2]')))
actions = ActionChains(driver)
actions.move_to_element(date_select)    # 選擇 2024.08.26 遊戲紀錄
actions.perform()
time.sleep(1)
date_select.click()

time.sleep(5)

# # =================================選取遊戲名稱 =================================

# gameList = WebDriverWait(driver, 5).until(
#     EC.element_to_be_clickable((By.XPATH, '//*[@id="game_select"]')))
# actions = ActionChains(driver)
# actions.move_to_element(gameList)    
# actions.perform()
# gameList.click()
# time.sleep(2)

# game_select = driver.find_element(By.XPATH, '//*[@id="game_select"]/option[148]')
# game_select.click()
# time.sleep(2)

# =======================================================寫入第一筆測試紀錄=======================================================

print('獲取前台遊戲紀錄中...', '\n')

xpath1 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]')

textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newList1 = DataFrame(newList1)
newList1T = newList1.T

csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826).csv', index=0, encoding='utf-8-sig')
readData = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826).csv')
newData = readData.rename(columns = {'0':'代理商', '1':'遊戲類型', '2':'紀錄流水號', '3':'代理伺服器', '4':'遊戲名稱',
                                     '5':'遊戲時間', '6':'場景D', '7':'日期', '8':'時間', '9':'場景', 
                                     '10':'序號D', '11':'押注D', '12':'序號', '13':'押注', '14':'彩金D',
                                     '15':'贏分D', '16':'彩金', '17':'贏分'})
newFile1 = newData.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', index=False, encoding = 'utf-8-sig')
print('第', 1, '筆資料寫入完成。', '\n')

for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i, '筆資料寫入完成。', '\n')     
    
# =================================選取遊戲紀錄頁(2) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[2]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.2。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+20, '筆資料寫入完成。', '\n')

# =================================選取遊戲紀錄頁(3) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[3]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.3。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+40, '筆資料寫入完成。', '\n') 

# =================================選取遊戲紀錄頁(4) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[4]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.4。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+60, '筆資料寫入完成。', '\n')
    
# =================================選取遊戲紀錄頁(5) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[5]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.5。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+80, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(6) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[6]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.6。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+100, '筆資料寫入完成。', '\n')

# =================================選取遊戲紀錄頁(7) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[7]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.7。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+120, '筆資料寫入完成。', '\n')

# =================================選取遊戲紀錄頁(8) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[8]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.8。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+140, '筆資料寫入完成。', '\n')

# =================================選取遊戲紀錄頁(9) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[9]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.9。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+160, '筆資料寫入完成。', '\n')

# =================================選取遊戲紀錄頁(10) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[10]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.10。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+180, '筆資料寫入完成。', '\n')

# =================================選取遊戲紀錄頁(11) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[11]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.11。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+200, '筆資料寫入完成。', '\n')   
   
# =================================選取遊戲紀錄頁(12) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[12]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.12。', '\n')
time.sleep(3)

for i in range(1, 31):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+230, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(13) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[13]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.13。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+250, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(14) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[14]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.14。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+270, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(15) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[15]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.15。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+290, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(16) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[16]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.16。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+310, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(17) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[17]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.17。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+330, '筆資料寫入完成。', '\n')

# =================================選取遊戲紀錄頁(18) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[18]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.18。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+350, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(19) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[19]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.19。', '\n')
time.sleep(3)

for i in range(1, 31):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+380, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(20) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[20]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.20。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+400, '筆資料寫入完成。', '\n')   

# =================================選取遊戲紀錄頁(21) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[21]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.21。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+420, '筆資料寫入完成。', '\n')   

# =================================選取遊戲紀錄頁(22) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[22]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.22。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+440, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(23) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[23]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.23。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+460, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(24) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[24]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.24。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    print('第', i+480, '筆資料寫入完成。', '\n')    

# =================================選取遊戲紀錄頁(25) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[25]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.25。', '\n')
time.sleep(3)

for i in range(1, 21):
    xpath = f'//*[@id="content"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    totalList = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv', on_bad_lines='skip')
    totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.xlsx', index=None, header=True)
    print('第', i+500, '筆資料寫入完成。', '\n')
    
print("EXCEL檔案儲存成功!", '\n')
time.sleep(2)   
driver.close()
print('(小豬大師)前台520筆遊戲紀錄獲取完成!', '\n')


# ============================Section.1 進入後台首頁============================

# ch_options = Options()
# ch_options.add_argument("--headless")  # => 無WEB UI顯示
# driver = webdriver.Chrome(ch_options)

driver = webdriver.Chrome()  # 有WEB UI顯示

print('串接DEV後台擷取遊戲紀錄測試中...', '\n')
time.sleep(1) 
driver.get("https://dev-admin-br-02.claretfox.com/")
http_status = requests.get('https://dev-admin-br-02.claretfox.com/')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print('Http Response Code:', stateCode, '\n')
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')
time.sleep(1)

back_platform = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div/div/div/form/div/div[1]")))
back_platform.click()
time.sleep(3)
print("進入DEV後台首頁!", '\n')

# -----------切換網頁顯示語系-----------
language_bar = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/form/label')))
language_bar.click()

language_ch_zh = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="lang"]/option[2]')))
language_ch_zh.click()
print("語系已切換'繁體中文'!", '\n')

# ============================Section.2 登入使用者名稱及密碼============================

login_icon = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, 'content-group')))
login_icon.click()  # 尋找登入介面元素位址
time.sleep(2)

login_acc = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="user_id"]')))
login_acc.clear()    # 預設此欄位為null, 但仍先清除帳號欄位資訊
login_pass = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]')))   
login_pass.clear()   # 預設此欄位為null, 但仍先清除密碼欄位資訊

login_verificationCode = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input')))  # 驗證碼欄位

verificationCode = driver.find_element(By.CSS_SELECTOR, '#captcha_img > img')
verificationCode.screenshot('verificationCode.png')
time.sleep(1)

actions = ActionChains(driver)
actions.move_to_element(login_acc)    
actions.perform()
login_acc.send_keys("ivan_li")   # 個人使用者帳號
time.sleep(1)
login_acc.send_keys(Keys.TAB)   # 切換至密碼輸入欄位
login_pass.send_keys("iPlaystar296")  # 個人密碼
time.sleep(1)
login_pass.send_keys(Keys.TAB)   # 切換至驗證碼輸入欄位
verifiCode = driver.find_element(By.XPATH, '//*[@id="captcha_img"]/img')  # 驗證碼圖片元素位置
data = verifiCode.screenshot_as_png
ocr = ddddocr.DdddOcr()
# with open('verificationCode.png', 'rb') as fp:
#     image = fp.read()
catch = ocr.classification(data)
login_verificationCode.send_keys(catch)
time.sleep(1)

try:
    login_button = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[6]/button')))
    login_button.click()
    time.sleep(1)
except Exception as e:
    # errorInfo = driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[7]')
    # # print('Error message = NoSuchElementException', '\n')
    # for errorCode in errorInfo:
        # print(errorCode, '\n')
    # 處理異常
    print("驗證碼輸入錯誤，請再試一次!", e)
    # print("驗證碼輸入錯誤，請再試一次!", '\n')
    driver.quit()  
else:
    print("登入成功!", '\n')

time.sleep(5)

# ============================Section.3 切換後台功能頁籤============================

player = driver.find_element(By.ID, 'Player')
player.click()
print("進入玩家功能選單!", '\n')
time.sleep(2)

driver.get('https://dev-admin-br-02.claretfox.com/Player/game_history')
print("切換遊戲紀錄子選單!", '\n')
time.sleep(2)

start_time = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div')))

startTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button')
actions.move_to_element(startTime_button)    
actions.perform()
startTime_button.click()

startDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(startDate)    
actions.perform()
startDate.click()

startDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(startDate)    
actions.perform()
startDate.click()

startTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[1]/a')  # 開始日期選取 "2024.08.26"
startTime_check.click()
time.sleep(1)

startTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
startTime_confirm.click()
time.sleep(2)

endTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i')
actions.move_to_element(endTime_button)    
actions.perform()
endTime_button.click()

endDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(endDate)    
actions.perform()
endDate.click()

endDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(endDate)    
actions.perform()
endDate.click()

endTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[1]/a')  # 結束日期選取 "2024.08.26"
endTime_check.click()
time.sleep(1)

endTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
endTime_confirm.click()
time.sleep(2)

showLimit = driver.find_element(By.XPATH, '//*[@id="count"]')
showLimit.click()
time.sleep(2)

showLimitSet = driver.find_element(By.XPATH, '//*[@id="count"]/option[6]')  # 設定顯示筆數 = 1000筆
showLimitSet.click()
time.sleep(2)

agencyType = driver.find_element(By.XPATH, '//*[@id="agent_attr"]')
agencyType.click()
time.sleep(2)

agencyTypeSet = driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[10]')  # 設定代理商類別 = Test
agencyTypeSet.click()
time.sleep(2)
print('代理商類別: ', agencyTypeSet.text, '\n')

agencyName = driver.find_element(By.XPATH, '//*[@id="agent"]')
agencyName.click()
time.sleep(2)

agencyNameSet = driver.find_element(By.XPATH, '//*[@id="agent"]/option[7]')  # 設定代理商名稱 = Test-2
agencyNameSet.click()
time.sleep(2)
print('代理商名稱: ', agencyNameSet.text, '\n')

gameType = driver.find_element(By.XPATH, '//*[@id="game_type"]')
gameType.click()
time.sleep(2)

gameTypeSet = driver.find_element(By.XPATH, '//*[@id="game_type"]/option[2]')  # 設定遊戲類別 = 老虎機
driver.execute_script("arguments[0].scrollIntoView()", gameTypeSet)
gameTypeSet.click()
time.sleep(2)
print('遊戲類別: ', gameTypeSet.text, '\n')

gameSelect = driver.find_element(By.XPATH, '//*[@id="game_select"]')
gameSelect.click()
time.sleep(2)

gameSelectSet = driver.find_element(By.XPATH, '//*[@id="game_select"]/option[159]')  # 遊戲選擇 = 小豬大師
gameSelectSet.click()
time.sleep(2)
print('遊戲名稱: ', gameSelectSet.text, '\n')

playerName = driver.find_element(By.XPATH, '//*[@id="player"]')
playerName.send_keys("ivan_li")
time.sleep(2)

btnSubmit = driver.find_element(By.XPATH, '//*[@id="sh_btn"]')  # 確認無誤送出
btnSubmit.click()
time.sleep(10)

print('獲取後台遊戲紀錄中...', '\n')

# =======================================================寫入第一筆測試紀錄=======================================================

xpath1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]')
textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newList1 = DataFrame(newList1)
newList1T = newList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826).csv', index=0, encoding='utf-8-sig')
readData = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826).csv')
newData = readData.rename(columns = { "0" : '紀錄流水號', "1" : '遊戲名稱', "2" : '遊戲編號', "3" : '玩家D', "4" : '代理伺服器', "5" : '玩家名稱', 
                                     "6" :'起始時間D', "7" :'結束時間D', "8" :'起始日期', "9" :'起始時間', "10" :'結束日期', 
                                     "11" :'結束時間', "12" :'遊戲狀態D', "13" :'場景D', "14" :'遊戲狀態', "15" :'場景', 
                                     "16" :'序號D', "17" :'面額D', "18" : '帳戶D', "19" : '押注D', "20" : '序號',
                                     "21" : '面額', "22" : '帳戶', "23" : '押注', "24" : '連線彩金D', "25" : '贏分D',
                                     "26" : '連線彩金', "27" : '贏分'})
newFile1 = newData.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826)_v1.csv', index=None, encoding='utf-8-sig')
print("第", 1, "筆資料寫入完成。", '\n')

# =======================================================寫入第2~520筆測試紀錄=======================================================

for i in range(2, 521):
    xpath = f'//*[@id="history"]/div[{i}]'
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
    newFile2 = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826)_v1.csv', on_bad_lines='skip')
    totalListEx = newFile2.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826)_v1.xlsx', index=None, header=True)
    print('第', i, '筆資料寫入完成。', '\n')

print("EXCEL檔案儲存成功!", '\n')
time.sleep(1)   
driver.close()
print('(小豬大師)後台520筆遊戲紀錄獲取完成!', '\n')

# =======================================================(前台)資料處理=======================================================

print('(前台)遊戲紀錄格式化...', '\n')

# 打开Excel文件
workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.xlsx')

# 选择要复制的工作表
source_sheet = workbook['Sheet1']

source_sheet.move_range('P223' ":" 'P232', rows = 0, cols = 2)
source_sheet.move_range('P383' ":" 'P392', rows = 0, cols = 2)

source_sheet.move_range('O223' ":" 'O232', rows = 0, cols = 1)
source_sheet.move_range('O383' ":" 'O392', rows = 0, cols = 1)
workbook.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1x.xlsx')
workbook.close()

df = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1x.xlsx').fillna(0)
df = df.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)
# print('Step.1 (前台)遊戲紀錄格式清理完成(空值=0)!', '\n')

df01 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1x.xlsx').drop(
{'代理商', '遊戲類型', '紀錄流水號', '遊戲時間', '場景D', '序號D', '押注D', '彩金D', '贏分D'}, axis=1)  # 刪除多餘欄位
df01 = df01.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)
# print('Step.2 (前台)遊戲紀錄格式清理完成(刪除多餘欄位)!', '\n')
print('(前台)遊戲紀錄格式化完成!', '\n')
time.sleep(1)

# =======================================================(後台)資料處理=======================================================

print('(後台)遊戲紀錄格式化...', '\n')

# 打开Excel文件
workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826)_v1.xlsx')

# 选择要复制的工作表
source_sheet = workbook['Sheet1']

source_sheet.move_range('W223' ":" 'W232', rows = 0, cols = 5)
source_sheet.move_range('W383' ":" 'W392', rows = 0, cols = 5)

source_sheet.move_range('V223' ":" 'V232', rows = 0, cols = 4)
source_sheet.move_range('V383' ":" 'V392', rows = 0, cols = 4)

source_sheet.move_range('U223' ":" 'U232', rows = 0, cols = 3)
source_sheet.move_range('U383' ":" 'U392', rows = 0, cols = 3)

source_sheet.move_range('T223' ":" 'T232', rows = 0, cols = 3)
source_sheet.move_range('T383' ":" 'T392', rows = 0, cols = 3)

source_sheet.move_range('S223' ":" 'S232', rows = 0, cols = 3)
source_sheet.move_range('S383' ":" 'S392', rows = 0, cols = 3)

source_sheet.move_range('R223' ":" 'R232', rows = 0, cols = 3)
source_sheet.move_range('R383' ":" 'R392', rows = 0, cols = 3)

source_sheet.move_range('P223' ":" 'P232', rows = 0, cols = 3)
source_sheet.move_range('P383' ":" 'P392', rows = 0, cols = 3)

source_sheet.move_range('M223' ":" 'M232', rows = 0, cols = 3)
source_sheet.move_range('M383' ":" 'M392', rows = 0, cols = 3)

source_sheet.move_range('J223' ":" 'J232', rows = 0, cols = 3)
source_sheet.move_range('J383' ":" 'J392', rows = 0, cols = 3)

source_sheet.move_range('I223' ":" 'I232', rows = 0, cols = 3)
source_sheet.move_range('I383' ":" 'I392', rows = 0, cols = 3)

source_sheet.move_range('H223' ":" 'H232', rows = 0, cols = 3)
source_sheet.move_range('H383' ":" 'H392', rows = 0, cols = 3)

workbook.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')
workbook.close()

df01 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx').fillna(0)  # 修改空值欄位=0
df01.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)
# print('Step.1 (後台)遊戲紀錄清理完成(空值=0)!', '\n')

df02 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx').drop(
{'紀錄流水號', '遊戲編號', '玩家D', '玩家名稱', '起始時間D', '結束時間D',
 '起始日期', '起始時間', '遊戲狀態D', '場景D', '遊戲狀態', '序號D',
 '面額D', '帳戶D', '押注D', '面額', '帳戶', '連線彩金D', '贏分D'}, axis=1)  # 刪除多餘欄位
df02.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)
# print('Step.2 (後台)遊戲紀錄清理完成(刪除多餘欄位)!', '\n')

df03 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx').rename(columns={'結束日期' : '日期', '結束時間' : '時間', '連線彩金' : '彩金'})
df03.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)  # 更改欄位名稱
# print('Step.3 (後台)遊戲紀錄清理完成(更新欄位名稱)!', '\n')

df04 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')
df04['代理伺服器'] = 'Test-2'
df04.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)  # 更改欄位值
# print('Step.4 (後台)遊戲紀錄清理完成(更新欄位值)!', '\n')

df05 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')
df05['彩金'] = 0
df05.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)  # 更改欄位值

workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')
source_sheet = workbook['Sheet1']
source_sheet.insert_cols(1, 1)
workbook.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')
# print('Step.5 (後台)遊戲紀錄更新完成(插入新欄位)!', '\n')

workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')
source_sheet = workbook['Sheet1']
source_sheet.move_range('C1' ':' 'C519', rows = 0, cols = -2)
source_sheet.delete_cols(3)
workbook.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')
# print('Step.6 (後台)遊戲紀錄更新完成(刪除欄位)!', '\n')
print('(後台)遊戲紀錄格式化完成!', '\n')
time.sleep(1)

# =================================================== Step.3 前/後台資料分析 ===================================================

print('前/後台遊戲紀錄比對中...', '\n')

# 1.數據比對

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1x.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1xa.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1xa.xlsx')
workbook_2.close()
time.sleep(1)
print('前/後台遊戲紀錄比對完成!', '\n')

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1xa.xlsx')
files.append(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1xa.xlsx')

print('前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xa_20240826.xlsx")
print("前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = "C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xa_20240826.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '前台遊戲紀錄比對結果(小豬大師)_20240826'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '後台遊戲紀錄比對結果(小豬大師)_20240826'  # 修改分頁.2工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xa_20240826.xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
wb.save("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xa_20240826.xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()
 
print('(小豬大師)前/後台共520筆遊戲紀錄擷取(含數據比對)測試完成!', '\n')

# =================================前台(免遊前)遊戲記錄擷取 =================================

print('串接(小豬大師)前台API擷取(免遊前)遊戲紀錄...', '\n')
time.sleep(1)

# ch_options = Options()
# ch_options.add_argument("--headless")  # 無WEB UI顯示
# driver = webdriver.Chrome(ch_options)
 
driver = webdriver.Chrome()   
driver.get('https://dev-api.iplaystar.net/gamehistory/?host_id=aa62ffb88b300f6be6654615f17ce6fa&lang=tch&game_id=PSS-ON-00157&count=20&page=1&uid=uN22JyloR1aD2zgUIu8nL1ogWFBPqrujwHHpugHyw94%3d')
http_status = requests.get('https://dev-api.iplaystar.net/gamehistory/?host_id=aa62ffb88b300f6be6654615f17ce6fa&lang=tch&game_id=PSS-ON-00157&count=20&page=1&uid=uN22JyloR1aD2zgUIu8nL1ogWFBPqrujwHHpugHyw94%3d')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print('Http Response Code:', stateCode, '\n')
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')
time.sleep(1)

# =================================選取記錄日期 =================================

dateList = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="serdate"]')))
actions = ActionChains(driver)
actions.move_to_element(dateList)    
actions.perform()
dateList.click()
time.sleep(2)

recordDate = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/thead/tr[2]/th[1]')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(recordDate)    
actions.perform()
recordDate.click()

date_select01 = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/thead/tr[2]/th[1]')
date_select01.click()
time.sleep(1)

date_select = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[7]/div[1]/table/tbody/tr[5]/td[2]')))
actions = ActionChains(driver)
actions.move_to_element(date_select)    # 選擇 2024.08.26 遊戲紀錄
actions.perform()
time.sleep(1)
date_select.click()

time.sleep(5)

# # =================================選取遊戲名稱 =================================

# gameList = WebDriverWait(driver, 5).until(
#     EC.element_to_be_clickable((By.XPATH, '//*[@id="game_select"]')))
# actions = ActionChains(driver)
# actions.move_to_element(gameList)    
# actions.perform()
# gameList.click()
# time.sleep(2)

# game_select = driver.find_element(By.XPATH, '//*[@id="game_select"]/option[147]')  # 遊戲名稱 = 小豬大師
# game_select.click()
# time.sleep(2)

# ================================= 取得(前台)免遊前遊戲紀錄(P.12) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[12]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.12。', '\n')
time.sleep(3)

print('獲取前台遊戲紀錄中...', '\n')

xpath1 = driver.find_element(By.XPATH, '//*[@id="content"]/div[12]')

textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newList1 = DataFrame(newList1)
newList1T = newList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826).csv', index=0, encoding='utf-8-sig')
readData = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826).csv')
newData = readData.rename(columns = {'0':'代理商', '1':'遊戲類型', '2':'紀錄流水號', '3':'代理伺服器', '4':'遊戲名稱',
                                     '5':'遊戲時間', '6':'場景D', '7':'日期', '8':'時間', '9':'場景', 
                                     '10':'序號D', '11':'押注D', '12':'序號', '13':'押注', '14':'彩金D',
                                     '15':'贏分D', '16':'彩金', '17':'贏分', '18':'免遊贏分'})
newFile1 = newData.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826)_v1.csv', index=False, encoding = 'utf-8-sig')
print('(前台)第', 1, '筆進入(免遊前)資料寫入完成。', '\n')

# ================================= 取得(前台)免遊前遊戲紀錄(P.19) =================================

listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[19]')
page_select.click()
print('已切換(小豬大師)_20240826遊戲紀錄頁.19。', '\n')
time.sleep(3)

xpath2 = driver.find_element(By.XPATH, '//*[@id="content"]/div[23]')

textSplit2 = xpath2.text.split()
newList2 = list(textSplit2)
newList2 = DataFrame(newList2)
newList2T = newList2.T
totalList = newList2T.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
print("CSV檔案匯出成功!", '\n')
totalList = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826)_v1.xlsx', index=False, header=True)
print("EXCEL檔案匯出成功!", '\n')
print('(前台)第', 2, '筆進入(免遊前)資料寫入完成。', '\n')
time.sleep(1)
driver.close()

# =======================================================(前台)進入免遊前紀錄資料處理=======================================================

print('(前台)進入(免遊前一筆)遊戲紀錄格式化...', '\n')

# 打开Excel文件
workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826)_v1.xlsx')

# 选择要复制的工作表
source_sheet = workbook['Sheet1']

source_sheet.move_range('R2' ":" 'R3', rows = 0, cols = -1)
source_sheet.move_range('S2' ":" 'S3', rows = 0, cols = -1)
source_sheet.move_range('T2' ":" 'T3', rows = 0, cols = -1)
source_sheet.delete_cols(20)
workbook.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')  

df1 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')
df2 = df1.drop({'代理商', '遊戲類型', '紀錄流水號', '遊戲時間', '場景D', '序號D', '押注D', '彩金D', '贏分D'}, axis=1)  # 刪除多餘欄位
df2.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_freerecord(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)
print("EXCEL檔案匯出成功!", '\n')
# print('(前台)遊戲紀錄清理完成(刪除多餘欄位)!', '\n')
print('(前台)進入(免遊前一筆)遊戲紀錄格式化完成!', '\n')
time.sleep(1)

# ============================Section.1 進入後台首頁============================

# ch_options = Options()
# ch_options.add_argument("--headless")  #無WEB UI顯示
# driver = webdriver.Chrome(ch_options)

driver = webdriver.Chrome()  # 有WEB UI顯示

print('串接DEV後台擷取(免遊前)遊戲紀錄...', '\n')
time.sleep(1) 
driver.get("https://dev-admin-br-02.claretfox.com/")
http_status = requests.get('https://dev-admin-br-02.claretfox.com/')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print('Http Response Code:', stateCode, '\n')
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')
time.sleep(2)

back_platform = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div/div/div/form/div/div[1]")))
back_platform.click()
time.sleep(2)
print("進入DEV後台首頁!", '\n')

# -----------切換網頁顯示語系-----------
language_bar = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/form/label')))
language_bar.click()

language_ch_zh = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="lang"]/option[2]')))
language_ch_zh.click()
print("語系已切換'繁體中文'!", '\n')

# ============================Section.2 登入使用者名稱及密碼============================

login_icon = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, 'content-group')))
login_icon.click()  # 尋找登入介面元素位址
time.sleep(2)

login_acc = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="user_id"]')))
login_acc.clear()    # 預設此欄位為null, 但仍先清除帳號欄位資訊
login_pass = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]')))   
login_pass.clear()   # 預設此欄位為null, 但仍先清除密碼欄位資訊

login_verificationCode = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input')))  # 驗證碼欄位

verificationCode = driver.find_element(By.CSS_SELECTOR, '#captcha_img > img')
verificationCode.screenshot('verificationCode.png')
time.sleep(1)

actions = ActionChains(driver)
actions.move_to_element(login_acc)    
actions.perform()
login_acc.send_keys("ivan_li")   # 個人使用者帳號
time.sleep(1)
login_acc.send_keys(Keys.TAB)   # 切換至密碼輸入欄位
login_pass.send_keys("iPlaystar296")  # 個人密碼
time.sleep(1)
login_pass.send_keys(Keys.TAB)   # 切換至驗證碼輸入欄位
verifiCode = driver.find_element(By.XPATH, '//*[@id="captcha_img"]/img')  # 驗證碼圖片元素位置
data = verifiCode.screenshot_as_png
ocr = ddddocr.DdddOcr()
# with open('verificationCode.png', 'rb') as fp:
#     image = fp.read()
catch = ocr.classification(data)
login_verificationCode.send_keys(catch)
time.sleep(1)

try:
    login_button = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[6]/button')))
    login_button.click()
    time.sleep(1)
except Exception as e:
    # errorInfo = driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[7]')
    # # print('Error message = NoSuchElementException', '\n')
    # for errorCode in errorInfo:
        # print(errorCode, '\n')
    # 處理異常
    print("驗證碼輸入錯誤，請再試一次!", e)
    # print("驗證碼輸入錯誤，請再試一次!", '\n')
    driver.quit()  
else:
    print("登入成功!", '\n')

time.sleep(5)

# ============================Section.3 切換後台功能頁籤============================

player = driver.find_element(By.ID, 'Player')
player.click()
print("進入玩家功能選單!", '\n')
time.sleep(2)

driver.get('https://dev-admin-br-02.claretfox.com/Player/game_history')
print("切換遊戲紀錄子選單!", '\n')
time.sleep(2)

start_time = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div')))

startTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button')
actions.move_to_element(startTime_button)    
actions.perform()
startTime_button.click()

startDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(startDate)    
actions.perform()
startDate.click()

startDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(startDate)    
actions.perform()
startDate.click()

startTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[1]/a')  # 開始日期選取 "2024.08.26"
startTime_check.click()
time.sleep(1)

startTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
startTime_confirm.click()
time.sleep(2)

endTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i')
actions.move_to_element(endTime_button)    
actions.perform()
endTime_button.click()

endDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(endDate)    
actions.perform()
endDate.click()

endDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(endDate)    
actions.perform()
endDate.click()

endTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[1]/a')  # 結束日期選取 "2024.08.26"
endTime_check.click()
time.sleep(1)

endTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
endTime_confirm.click()
time.sleep(2)

showLimit = driver.find_element(By.XPATH, '//*[@id="count"]')
showLimit.click()
time.sleep(2)

showLimitSet = driver.find_element(By.XPATH, '//*[@id="count"]/option[6]')  # 設定顯示筆數 = 1000筆
showLimitSet.click()
time.sleep(2)

agencyType = driver.find_element(By.XPATH, '//*[@id="agent_attr"]')
agencyType.click()
time.sleep(2)

agencyTypeSet = driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[10]')  # 設定代理商類別 = Test
agencyTypeSet.click()
time.sleep(2)
print('代理商類別: ', agencyTypeSet.text, '\n')

agencyName = driver.find_element(By.XPATH, '//*[@id="agent"]')
agencyName.click()
time.sleep(2)

agencyNameSet = driver.find_element(By.XPATH, '//*[@id="agent"]/option[7]')  # 設定代理商名稱 = Test-2
agencyNameSet.click()
time.sleep(2)
print('代理商名稱: ', agencyNameSet.text, '\n')

gameType = driver.find_element(By.XPATH, '//*[@id="game_type"]')
gameType.click()
time.sleep(2)

gameTypeSet = driver.find_element(By.XPATH, '//*[@id="game_type"]/option[2]')  # 設定遊戲類別 = 老虎機
driver.execute_script("arguments[0].scrollIntoView()", gameTypeSet)
gameTypeSet.click()
time.sleep(2)
print('遊戲類別: ', gameTypeSet.text, '\n')

gameSelect = driver.find_element(By.XPATH, '//*[@id="game_select"]')
gameSelect.click()
time.sleep(2)

gameSelectSet = driver.find_element(By.XPATH, '//*[@id="game_select"]/option[159]')  # 遊戲選擇 = 小豬大師
gameSelectSet.click()
time.sleep(2)
print('遊戲名稱: ', gameSelectSet.text, '\n')

playerName = driver.find_element(By.XPATH, '//*[@id="player"]')
playerName.send_keys("ivan_li")
time.sleep(2)

btnSubmit = driver.find_element(By.XPATH, '//*[@id="sh_btn"]')  # 確認無誤送出
btnSubmit.click()
time.sleep(10)

print('獲取後台遊戲紀錄中...', '\n')

# ================================= 取得(後台)免遊前遊戲紀錄 =================================

xpath1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[232]')

textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)

newList1 = DataFrame(newList1)
newList1T = newList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826).csv', index=0, encoding='utf-8-sig')
readData = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826).csv')

newData = readData.rename(columns = { "0" : '紀錄流水號', "1" : '遊戲名稱', "2" : '遊戲編號', "3" : '玩家D', "4" : '代理伺服器', "5" : '玩家名稱', 
                                     "6" :'起始時間D', "7" :'結束時間D', "8" :'起始日期', "9" :'起始時間', "10" :'結束日期', 
                                     "11" :'結束時間', "12" :'遊戲狀態D', "13" :'場景D', "14" :'遊戲狀態', "15" :'場景', 
                                     "16" :'序號D', "17" :'面額D', "18" : '帳戶D', "19" : '押注D', "20" : '序號',
                                     "21" : '面額', "22" : '帳戶', "23" : '押注', "24" : '彩金D', "25" : '贏分D',
                                     "26" : '免遊贏分D', "27" : '彩金', "28" : '贏分', "29" : '免遊贏分'})
newFile1 = newData.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.csv', index=None, encoding='utf-8-sig')
totalList = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.xlsx', index=False, header=True)
print("(後台)第", 1, "筆進入(免遊前)資料寫入完成。", '\n')

xpath2 = driver.find_element(By.XPATH, '//*[@id="history"]/div[393]')

textSplit2 = xpath2.text.split()
newList2 = list(textSplit2)
newList2 = DataFrame(newList2)
newList2T = newList2.T
totalList = newList2T.to_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.csv', mode='a', header=False, index=None, encoding='utf-8-sig')
print("CSV檔案匯出成功!", '\n')
# print('檔案路徑= ', 'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.csv', '\n')
time.sleep(1)
totalList = pd.read_csv(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.csv')
totalListEx = totalList.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.xlsx', index=False, header=True)
print("EXCEL檔案匯出成功!", '\n')
print('(後台)第', 2, '筆進入(免遊前)資料寫入完成。', '\n')
time.sleep(1)
driver.close()

# =======================================================(後台)進入免遊前紀錄資料處理=======================================================

print('(後台)進入(免遊前一筆)遊戲紀錄格式化...', '\n')

df1 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.xlsx')
df2 = df1.drop({'紀錄流水號', '遊戲編號', '玩家D', '玩家名稱', '起始時間D', '結束時間D',
 '起始日期', '起始時間', '遊戲狀態D', '場景D', '遊戲狀態', '序號D',
 '面額D', '帳戶D', '押注D', '面額', '帳戶', '彩金D', '贏分D', '免遊贏分D'}, axis=1)  # 刪除多餘欄位
df2.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)
# print('(後台)遊戲紀錄更新完成(刪除多餘欄位)!', '\n')

df3 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx').rename(columns={'結束日期' : '日期', '結束時間' : '時間'})
df3.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)  # 更改欄位名稱
# print('(後台)遊戲紀錄更新完成(更改欄位名稱)!', '\n')

df4 = pd.read_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')
df4['代理伺服器'] = 'Test-2'
df4.to_excel(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx', index=None, header=True)  # 更改欄位值
# print('(後台)遊戲紀錄更新完成(更改欄位值)!', '\n')

workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')
source_sheet = workbook['Sheet1']
source_sheet.insert_cols(1, 1)
workbook.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')
# print('(後台)遊戲紀錄更新完成(插入新欄位)!', '\n')

workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')
source_sheet = workbook['Sheet1']
source_sheet.move_range('C1' ':' 'C3', rows = 0, cols = -2)
source_sheet.delete_cols(3)
workbook.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')
print("EXCEL檔案匯出成功!", '\n')
# print('(後台)遊戲紀錄更新完成(欄位資料搬移)!', '\n')
print('(後台)進入(免遊前一筆)遊戲紀錄格式化完成!', '\n')
time.sleep(1)

# =================================================== Step.3 前/後台資料處理 ===================================================

print('前/後台(免遊前一筆)遊戲紀錄比對中...', '\n')

# 1.數據比對

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 最后将对比处理完成后的工作表对象使用save函数进行保存即可。

# 将 前/後台有差異欄位標記"黃底"並匯出
print('前台(免遊前一筆)比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_freerecord(Pig-Master_20240826)_v1xa.xlsx')
print('前台(免遊前一筆)比對資料匯出成功!', '\n')
workbook_1.close()
print('後台(免遊前一筆)比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1xa.xlsx')
print('後台(免遊前一筆)比對資料匯出成功!', '\n')
workbook_2.close()
print('前/後台(免遊前一筆)遊戲紀錄比對完成!', '\n')

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_freerecord(Pig-Master_20240826)_v1xa.xlsx')
files.append(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1xa.xlsx')
print('前/後台(免遊前一筆)比對紀錄合併中...', '\n') 
 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xb_20240826.xlsx")
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = "C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xb_20240826.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '前台(進入免遊前一筆)遊戲紀錄比對結果(小豬大師)_20240826'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '後台(進入免遊前一筆)遊戲紀錄比對結果(小豬大師)_20240826'  # 修改分頁.2工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('前/後台(進入免遊前一筆)比對紀錄合併完成!', '\n')
time.sleep(1)

print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xb_20240826.xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
wb.save("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xb_20240826.xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xa_20240826.xlsx')
files.append(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xb_20240826.xlsx')
print('前/後台比對紀錄(含進入免遊前一筆)合併中...', '\n') 
 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xc_20240826.xlsx")
newbook.Dispose()
tempbook.Dispose()

filename = "C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xc_20240826.xlsx"
wb = openpyxl.load_workbook(filename)

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('前/後台比對紀錄(含進入免遊前一筆)合併完成!', '\n')
time.sleep(1)

print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xc_20240826.xlsx")
sheet_1 = wb.worksheets[2]  # 分頁.3
sheet_1.sheet_properties.tabColor = '886600'
sheet_2 = wb.worksheets[3]  # 分頁.4
sheet_2.sheet_properties.tabColor = 'FFCC22'
wb.save("C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xc_20240826.xlsx")
# print('檔案路徑= ', r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xc_20240826.xlsx', '\n')
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()

# =========================================================== 刪除多餘資料 ===========================================================

os.remove(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826).csv')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826).csv')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826).csv')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826).csv')

os.remove(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Pig-Master_20240826)_v1.csv')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Pig-Master_20240826)_v1.csv')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_freerecord(Pig-Master_20240826)_v1.csv')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_freerecord(Pig-Master_20240826)_v1.csv')

os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1x.xlsx')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1x.xlsx')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1x.xlsx')

os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_record(Pig-Master_20240826)_v1xa.xlsx')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_record(Pig-Master_20240826)_v1xa.xlsx')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(1). Front-platform\front-platform_freerecord(Pig-Master_20240826)_v1xa.xlsx')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(2). Back-platform\back-platform_freerecord(Pig-Master_20240826)_v1xa.xlsx')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xa_20240826.xlsx')
os.remove(r'C:\AutomotiveTest\(2). Pig Master\(3). Data Analysis\(3). Data_Merge\Data Merge_v1xb_20240826.xlsx') 

 
print('(小豬大師)前/後台遊戲紀錄擷取(含自動化數據分析)測試完成!', '\n')
time.sleep(1) 
 
end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
print('測試結束時間: ', end_time, '\n')

# 取得現在時間
now = datetime.datetime.now()
txt = '上次執行時間為：' + str(now)

# 轉成df
dfLog = pd.DataFrame([txt], index=['UpdateTime'])

# 存出檔案
dfLog.to_csv(r'C:\AutomotiveTest\(2). Pig Master\log.csv', header=False)
